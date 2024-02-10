import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook, Workbook

DEFAULT_LOCATION_FILE = "default_location.txt"


class GUIApp:
    def __init__(self, root):
        self.root = root
        self.root.title("주소 입력 GUI")

        # 통신사 체크, 통신사 예외 처리용 변수 생성 kt 인지 sk 인지 lg 인지 확인용
        self.check_landlord_communication = {"comm": ""}
        self.check_tenant_communication = {"comm": ""}
        self.check_realtor_communication = {"comm": ""}

        self.previous_values = {
            "area": "",
            "dong": "",
            "number": "",
            "landlord_phone": "",
            "tenant_phone": "",
            "realtor_phone": "",
            "file_path": "",
        }

        # 지역, 동, 지번 입력을 위한 레이블 및 입력 상자 생성
        self.area_label = tk.Label(root, text="지역")
        self.area_label.grid(row=0, column=0, columnspan=30, padx=5, pady=5, sticky="w")

        # IntVar를 사용하여 라디오 버튼 값을 저장
        self.area_var = tk.IntVar()

        # 라디오 버튼 생성
        tk.Radiobutton(root, text="성동구", variable=self.area_var, value=1).grid(
            row=0, column=1, columnspan=30, padx=5, pady=5, sticky="w"
        )
        tk.Radiobutton(root, text="강남구", variable=self.area_var, value=2).grid(
            row=0, column=3, columnspan=30, padx=5, pady=5, sticky="w"
        )
        tk.Radiobutton(root, text="서초구", variable=self.area_var, value=3).grid(
            row=0, column=5, columnspan=30, padx=5, pady=5, sticky="w"
        )

        # 텍스트 상자 입력 기능 -> 체크박스로 변경(entry에는 체크 박스 기능 후 문자열 입력)
        self.area_entry = ""
        # self.area_entry = tk.Entry(root, width=30)
        # self.area_entry.grid(row=0, column=1, columnspan=30, padx=5, pady=5, sticky="w")

        self.dong_label = tk.Label(root, text="동")
        self.dong_label.grid(row=1, column=0, columnspan=30, padx=5, pady=5, sticky="w")
        self.dong_entry = tk.Entry(root, width=30)
        self.dong_entry.grid(row=1, column=1, columnspan=30, padx=5, pady=5, sticky="w")

        self.number_label = tk.Label(root, text="지번")
        self.number_label.grid(
            row=2, column=0, columnspan=30, padx=5, pady=5, sticky="w"
        )
        self.number_entry = tk.Entry(root, width=30)
        self.number_entry.grid(
            row=2, column=1, columnspan=30, padx=5, pady=5, sticky="w"
        )

        # 임대인, 임차인, 부동산 전화번호 입력을 위한 레이블 및 입력 상자 생성
        self.landlord_phone_labels = [
            tk.Label(root, text="임대인 전화번호"),
            tk.Label(root, text="-"),
            tk.Label(root, text="-"),
        ]

        self.tenant_phone_labels = [
            tk.Label(root, text="임차인 전화번호"),
            tk.Label(root, text="-"),
            tk.Label(root, text="-"),
        ]
        self.realtor_phone_labels = [
            tk.Label(root, text="부동산 전화번호"),
            tk.Label(root, text="-"),
            tk.Label(root, text="-"),
        ]

        self.landlord_phone_entries = [
            tk.Entry(root, width=6),
            tk.Entry(root, width=6),
            tk.Entry(root, width=6),
        ]
        self.tenant_phone_entries = [
            tk.Entry(root, width=6),
            tk.Entry(root, width=6),
            tk.Entry(root, width=6),
        ]
        self.realtor_phone_entries = [
            tk.Entry(root, width=6),
            tk.Entry(root, width=6),
            tk.Entry(root, width=6),
        ]

        for i, label in enumerate(self.landlord_phone_labels):
            label.grid(row=3, column=i * 2, padx=5, pady=5, sticky="w")
        for i, entry in enumerate(self.landlord_phone_entries):
            entry.grid(row=3, column=i * 2 + 1, padx=5, pady=5, sticky="w")

        for i, label in enumerate(self.tenant_phone_labels):
            label.grid(row=4, column=i * 2, padx=5, pady=5, sticky="w")
        for i, entry in enumerate(self.tenant_phone_entries):
            entry.grid(row=4, column=i * 2 + 1, padx=5, pady=5, sticky="w")

        for i, label in enumerate(self.realtor_phone_labels):
            label.grid(row=5, column=i * 2, padx=5, pady=5, sticky="w")
        for i, entry in enumerate(self.realtor_phone_entries):
            entry.grid(row=5, column=i * 2 + 1, padx=5, pady=5, sticky="w")

        # 체크박스 추가
        self.allow_blank_landlord = tk.BooleanVar()
        self.allow_blank_tenant = tk.BooleanVar()
        self.allow_blank_realtor = tk.BooleanVar()

        tk.Checkbutton(
            root,
            text="임대인 전화번호 공백 허용",
            variable=self.allow_blank_landlord,
        ).grid(row=3, column=7, padx=5, pady=5, sticky="w")

        tk.Checkbutton(
            root,
            text="임차인 전화번호 공백 허용",
            variable=self.allow_blank_tenant,
        ).grid(row=4, column=7, padx=5, pady=5, sticky="w")

        tk.Checkbutton(
            root,
            text="부동산 전화번호 공백 허용",
            variable=self.allow_blank_realtor,
        ).grid(row=5, column=7, padx=5, pady=5, sticky="w")

        # 파일 경로 입력을 위한 레이블 및 입력 상자 생성
        self.file_frame = tk.Frame(root)
        self.file_frame.grid(row=6, column=0, columnspan=30, padx=5, pady=5, sticky="w")

        self.location_label = tk.Label(self.file_frame, text="저장 위치")
        self.location_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")

        self.file_entry = tk.Entry(self.file_frame, width=30)
        self.file_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.file_entry.insert(0, self.previous_values["file_path"])

        self.browse_button = tk.Button(
            self.file_frame, text="찾아보기", command=self.browse_file
        )
        self.browse_button.grid(row=0, column=2, padx=5, pady=5, sticky="w")

        self.default_location_button = tk.Button(
            self.file_frame, text="기본 위치 지정", command=self.set_default_location
        )
        self.default_location_button.grid(row=0, column=3, padx=5, pady=5, sticky="w")

        # 프로그램 실행 시 기본 저장 위치를 로드
        self.load_default_location()

        # 작업 기록 출력을 위한 텍스트 박스와 스크롤 바 생성
        self.log_text_frame = tk.Frame(root)
        self.log_text_frame.grid(
            row=7, column=0, columnspan=30, padx=5, pady=5, sticky="w"
        )

        self.log_text = tk.Text(
            self.log_text_frame, height=10, width=65, state=tk.DISABLED
        )
        self.log_text.pack(padx=10, pady=10, anchor="w")

        # 버튼 생성
        self.button_submit = tk.Button(root, text="제출", command=self.on_button_click)
        self.button_submit.grid(
            row=8, column=0, columnspan=5, padx=5, pady=5, sticky="e"
        )

        self.button_cancel = tk.Button(root, text="취소", command=self.on_button_cancel)
        self.button_cancel.grid(
            row=8, column=1, columnspan=6, padx=5, pady=5, sticky="e"
        )

        # 결과 메시지 출력 레이블
        self.label_result = tk.Label(root, text="")
        self.label_result.grid(
            row=9, column=0, columnspan=30, padx=5, pady=5, sticky="w"
        )

        # 엔터 키 이벤트 바인딩
        self.root.bind("<Return>", self.on_enter)

    def on_button_click(self):
        area_input = ""
        # reaa_input = self.area_entry.get()
        dong_input = self.dong_entry.get()
        number_input = self.number_entry.get()
        area_var = self.area_var  # 성동구

        # 지역 체크버튼 기능
        if area_var.get() == 1:
            area_input = "성동구"
        elif area_var.get() == 2:
            area_input = "강남구"
        elif area_var.get() == 3:
            area_input = "서초구"
        else:
            area_input = "error"

        # Validate phone numbers based on checkbox states
        # flag 0 : landlord, 1 : tenant, 2:realtor
        landlord_phone_input = (
            ""
            if self.allow_blank_landlord.get()
            else self.get_phone_number(self.landlord_phone_entries, 0)
        )

        tenant_phone_input = (
            ""
            if self.allow_blank_tenant.get()
            else self.get_phone_number(self.tenant_phone_entries, 1)
        )
        realtor_phone_input = (
            ""
            if self.allow_blank_realtor.get()
            else self.get_phone_number(self.realtor_phone_entries, 2)
        )

        # 입력값이 모두 있는지 확인
        if not area_input or not dong_input or not number_input:
            self.label_result.config(text="지역, 동, 지번은 필수 입력 항목입니다.")
            return

        file_path = self.file_entry.get()
        check_landlord_communication = self.check_landlord_communication["comm"]
        check_tenant_communication = self.check_tenant_communication["comm"]
        check_realtor_communication = self.check_realtor_communication["comm"]

        # 엑셀 파일에 입력값 추가
        self.create_or_update_excel(
            file_path,
            area_input,
            dong_input,
            number_input,
            check_landlord_communication,
            landlord_phone_input,
            check_tenant_communication,
            tenant_phone_input,
            check_realtor_communication,
            realtor_phone_input,
        )  # 여기에 입력 값 추가

        # 작업 내용 로그에 추가
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(
            tk.END, f"입력 값 추가: {area_input}, {dong_input}, {number_input}\n"
        )
        self.log_text.config(state=tk.DISABLED)

        # 입력 값 초기화

        self.area_entry = ""
        # self.area_entry.delete(0, tk.END)
        self.dong_entry.delete(0, tk.END)
        self.number_entry.delete(0, tk.END)
        for entry in (
            self.landlord_phone_entries
            + self.tenant_phone_entries
            + self.realtor_phone_entries
        ):
            entry.delete(0, tk.END)
        self.file_entry.delete(0, tk.END)

        # 해당 Checkbutton 체크 해제
        self.area_var.set(0)

        # 결과 메시지 표시
        self.label_result.config(text="입력이 성공적으로 완료되었습니다.")
        self.load_default_location()

    def on_button_cancel(self):
        # 이전 값으로 입력 상자 초기화
        self.area_entry = ""
        # self.area_entry.delete(0, tk.END)
        self.dong_entry.delete(0, tk.END)
        self.number_entry.delete(0, tk.END)
        for entry in (
            self.landlord_phone_entries
            + self.tenant_phone_entries
            + self.realtor_phone_entries
        ):
            entry.delete(0, tk.END)
        self.file_entry.delete(0, tk.END)

        # 해당 Checkbutton 체크 해제
        self.area_var.set(0)

        # 취소 메시지
        self.label_result.config(text="입력이 취소되었습니다.")
        self.load_default_location()

    """ 엑셀 내용 수정 코드(미완성)
    def remove_entry_from_excel(self, file_path, area, dong, number):
        try:
            if self.file_exists(file_path):
                wb = load_workbook(file_path)
                sheet = wb.active

                # 이전 입력 값 지우기
                for i in range(1, sheet.max_row + 1):
                    if (
                        sheet.cell(i, 1).value == area
                        and sheet.cell(i, 2).value == dong
                        and sheet.cell(i, 3).value == number
                    ):
                        sheet.delete_rows(i, 1)
                        wb.save(file_path)
                        self.log_text.config(state=tk.NORMAL)
                        self.log_text.insert(
                            tk.END, f"이전 값 삭제 성공: {area}, {dong}, {number}\n"
                        )
                        self.log_text.config(state=tk.DISABLED)
                        break
        except Exception as e:
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, f"이전 값 삭제 중 오류 발생: {e}\n")
            self.log_text.config(state=tk.DISABLED)
    """

    def on_enter(self, event):
        self.on_button_click()

    def browse_file(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")]
        )
        self.file_entry.delete(0, tk.END)
        self.file_entry.insert(0, file_path)

    def create_or_update_excel(
        self,
        file_path,
        area,
        dong,
        number,
        ch_landlord_p,
        landlord_phone,
        ch_tenant_p,
        tenant_phone,
        ch_realtor_p,
        realtor_phone,
    ):
        try:
            if not self.file_exists(file_path):
                self.create_excel(file_path)

            wb = load_workbook(file_path)
            sheet = wb.active

            """
            # 이전 입력 값 지우기
            for i in range(1, sheet.max_row + 1):
                if (
                    sheet.cell(i, 1).value == area
                    and sheet.cell(i, 2).value == dong
                    and sheet.cell(i, 3).value == number
                ):
                    sheet.delete_rows(i, 1)
                    break
            """
            # 각각의 값이 하나의 셀에 입력되도록 수정
            sheet.append(
                [
                    area,
                    dong,
                    number,
                    ch_landlord_p,
                    landlord_phone,
                    ch_tenant_p,
                    tenant_phone,
                    ch_realtor_p,
                    realtor_phone,
                ]
            )

            wb.save(file_path)
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, f"엑셀 파일 업데이트 성공: {file_path}\n")
            self.log_text.config(state=tk.DISABLED)

        except Exception as e:
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, f"엑셀 파일 업데이트 중 오류 발생: {e}\n")
            self.log_text.config(state=tk.DISABLED)

    def create_excel(self, file_path):
        wb = Workbook()
        sheet = wb.active

        # 헤더 추가
        sheet.append(
            [
                "지역",
                "동",
                "지번",
                "임대인 통신사",
                "임대인 전화번호",
                "임차인 통신사",
                "임차인 전화번호",
                "부동산 통신사",
                "부동산 전화번호",
            ]
        )

        wb.save(file_path)
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"새로운 엑셀 파일 생성: {file_path}\n")
        self.log_text.config(state=tk.DISABLED)

    def file_exists(self, file_path):
        try:
            with open(file_path, "r"):
                return True
        except FileNotFoundError:
            return False

    def set_default_location(self):
        previous_location = self.file_entry.get()
        self.file_entry.delete(0, tk.END)
        self.file_entry.insert(0, previous_location)

        # 기본 저장 위치를 텍스트 파일에 저장
        with open(DEFAULT_LOCATION_FILE, "w") as f:
            f.write(previous_location)

    def load_default_location(self):
        try:
            with open(DEFAULT_LOCATION_FILE, "r") as f:
                default_location = f.read()
                self.file_entry.delete(0, tk.END)
                self.file_entry.insert(0, default_location)
        except FileNotFoundError:
            pass

    def validate_phone_entry(self, entry, index, max_length):
        value = entry.get()
        if not value.isdigit():
            entry.delete(0, tk.END)
            entry.insert(0, value[:-1])
        if len(value) > max_length:
            entry.delete(0, tk.END)
            entry.insert(0, value[:max_length])

    def get_phone_number(self, entries, flag):
        # flag 0 : landlord, 1 : tenant, 2:realtor
        if flag == 0:  # landlord
            self.check_communication_entries(entries, self.check_landlord_communication)
        elif flag == 1:  # tenant
            self.check_communication_entries(entries, self.check_tenant_communication)
        else:  # realtor
            self.check_communication_entries(entries, self.check_realtor_communication)

        return "-".join(entry.get() for entry in entries if entry.get())

    def check_communication_entries(self, entries, check_comm):
        if int(entries[2].get()[0]) == 2:
            if int(entries[2].get()[1]) == 0:
                check_comm["comm"] = "SKT"
            elif 2 <= int(entries[2].get()[1]) and int(entries[2].get()[1]) <= 4:
                check_comm["comm"] = "LGU+"
            elif 5 <= int(entries[2].get()[1]):
                check_comm["comm"] = "KT"
            else:
                check_comm["comm"] = ""

        elif int(entries[2].get()[0]) == 3:
            if int(entries[2].get()[1]) == 0:
                check_comm["comm"] = "KT"
            elif int(entries[2].get()[1]) == 1:
                check_comm["comm"] = "SKT"
            elif 2 <= int(entries[2].get()[1]) and int(entries[2].get()[1]) <= 4:
                check_comm["comm"] = "KT"
            elif 5 <= int(entries[2].get()[1]) and int(entries[2].get()[1]) <= 8:
                check_comm["comm"] = "SKT"
            elif int(entries[2].get()[1]) == 9:
                check_comm["comm"] = "LGU+"
            else:
                check_comm["comm"] = ""

        elif int(entries[2].get()[0]) == 4:
            if int(entries[2].get()[1]) == 0:
                check_comm["comm"] = "SKT"
            elif int(entries[2].get()[1]) == 1:
                check_comm["comm"] = "SKT"
            elif 2 <= int(entries[2].get()[1]) and int(entries[2].get()[1]) <= 4:
                check_comm["comm"] = "KT"
            elif 5 <= int(entries[2].get()[1]):
                check_comm["comm"] = "SKT"
            else:
                check_comm["comm"] = ""

        elif int(entries[2].get()[0]) == 5:
            if int(entries[2].get()[1]) == 0:
                check_comm["comm"] = "SKT"
            elif int(entries[2].get()[1]) == 1:
                check_comm["comm"] = "KT"
            elif 2 <= int(entries[2].get()[1]) and int(entries[2].get()[1]) <= 4:
                check_comm["comm"] = "SKT"
            elif 5 <= int(entries[2].get()[1]) and int(entries[2].get()[1]) <= 8:
                check_comm["comm"] = "LGU+"
            else:
                check_comm["comm"] = ""

        elif int(entries[2].get()[0]) == 6:
            if 2 <= int(entries[2].get()[1]) and int(entries[2].get()[1]) <= 4:
                check_comm["comm"] = "SKT"
            elif 5 <= int(entries[2].get()[1]) and int(entries[2].get()[1]) <= 8:
                check_comm["comm"] = "KT"
            else:
                check_comm["comm"] = ""

        elif int(entries[2].get()[0]) == 7:
            if int(entries[2].get()[1]) == 1:
                check_comm["comm"] = "SKT"
            elif 2 <= int(entries[2].get()[1]) and int(entries[2].get()[1]) <= 4:
                check_comm["comm"] = "KT"
            elif 5 <= int(entries[2].get()[1]) and int(entries[2].get()[1]) <= 7:
                check_comm["comm"] = "LGU+"
            elif int(entries[2].get()[1]) == 9:
                check_comm["comm"] = "LGU+"
            else:
                check_comm["comm"] = ""

        elif int(entries[2].get()[0]) == 8:
            if 0 <= int(entries[2].get()[1]) and int(entries[2].get()[1]) <= 4:
                check_comm["comm"] = "LGU+"
            elif 5 <= int(entries[2].get()[1]) and int(entries[2].get()[1]) <= 9:
                check_comm["comm"] = "SKT"
            else:
                check_comm["comm"] = ""

        elif int(entries[2].get()[0]) == 9:
            if 0 <= int(entries[2].get()[1]) and int(entries[2].get()[1]) <= 4:
                check_comm["comm"] = "SKT"
            elif 5 <= int(entries[2].get()[1]) and int(entries[2].get()[1]) <= 9:
                check_comm["comm"] = "KT"
            else:
                check_comm["comm"] = ""


if __name__ == "__main__":
    root = tk.Tk()
    app = GUIApp(root)
    root.mainloop()
